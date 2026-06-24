from __future__ import annotations

import hashlib
import json
import os
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import HTTPException
from sqlalchemy import Date, DateTime, Enum, Integer, Numeric, delete, func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.config import REPO_ROOT
from app.domain.holidays import business_days_between, month_workday_summary
from app.domain.people import apply_personnel_scope
from app.enums import (
    AssignmentStatus,
    AssignmentType,
    EmploymentStatus,
    ProjectAssignmentRole,
    ProjectLogStatus,
    ProjectStatus,
    ProjectType,
    UserPermission,
)
from app.models.core import (
    CurrentAssignmentSnapshot,
    JobLock,
    MonthlyAssignmentMM,
    MonthlyEmploymentMM,
    MonthlyKpiSummary,
    Personnel,
    Project,
    ProjectAssignment,
    ProjectCode,
    ProjectLog,
    Role,
    User,
    WeeklyLoadSnapshot,
)

DATA_BACKUP_JOB_NAME = "data_backup_upload_restore"
VALIDATION_TTL_MINUTES = 30
DATA_RUNTIME_DIR = REPO_ROOT / "apps" / "api" / ".runtime" / "data_backup"
BACKUP_DIR = DATA_RUNTIME_DIR / "backups"
MANIFEST_PATH = DATA_RUNTIME_DIR / "manifest.json"
LOCK_EXPIRE_MINUTES = 30
INCLUDED_TABLE_NAMES = [
    "personnel",
    "project_codes",
    "projects",
    "project_assignments",
    "project_logs",
    "monthly_employment_mm",
    "monthly_assignment_mm",
    "current_assignment_snapshots",
    "weekly_load_snapshots",
    "monthly_kpi_summaries",
]
PROJECT_SHEET_NAME = "프로젝트관리"
PERSONNEL_SHEET_NAME = "인력관리"
ASSIGNMENT_SHEET_NAME = "프로젝트배정"

PROJECT_HEADERS = [
    "코드",
    "사업명",
    "고객사",
    "사업유형",
    "확도",
    "사업금액",
    "공고번호",
    "공고일",
    "상태",
    "영업부서",
    "영업대표",
    "제안PM",
    "발표PM",
    "수행PM",
    "제안/수행팀",
    "사업 시작일",
    "사업 종료일",
    "제안서 제출일",
    "제출 형식",
    "제출 유의사항",
    "제안 발표일",
    "발표 형식",
    "발표 유의사항",
    "최근 활동일",
    "사용여부",
]
PERSONNEL_HEADERS = [
    "사번",
    "성명",
    "이메일",
    "본부",
    "팀",
    "직위",
    "역할",
    "재직상태",
    "MM 시작일",
    "MM 종료일",
    "연간 재직 MM",
    "사용여부",
]
ASSIGNMENT_HEADERS = [
    "배정키",
    "프로젝트코드",
    "사번",
    "배정유형",
    "프로젝트 역할",
    "배정상태",
    "WIN/LOSS",
    "상주유형",
    "대표여부",
    "순서",
    "시작일",
    "종료일",
    "MM",
    "총 MM",
    "현재 MM",
    "확도율",
    "단가",
    "비고",
]
PROJECT_TYPE_MAP = {
    "주사업": ProjectType.MAIN,
    "부사업": ProjectType.SUB,
    "하도": ProjectType.SUBCONTRACT,
    "협력": ProjectType.PARTNER,
}
PROJECT_STATUS_MAP = {
    "제안중": ProjectStatus.PROPOSING,
    "발표완료": ProjectStatus.PRESENTED,
    "WIN": ProjectStatus.WIN,
    "LOSS": ProjectStatus.LOSS,
    "DROP": ProjectStatus.DROP,
    "수행중": ProjectStatus.RUNNING,
    "업무지원": ProjectStatus.SUPPORT,
    "완료": ProjectStatus.DONE,
}
EMPLOYMENT_STATUS_MAP = {
    "재직": EmploymentStatus.ACTIVE,
    "휴직": EmploymentStatus.LEAVE,
    "전배": EmploymentStatus.TRANSFERRED,
    "퇴직": EmploymentStatus.RETIRED,
    "대기": EmploymentStatus.WAITING,
}
ASSIGNMENT_TYPE_MAP = {
    "수행": AssignmentType.DELIVERY,
    "제안": AssignmentType.PROPOSAL,
    "지원": AssignmentType.SUPPORT,
    "대기": AssignmentType.UNASSIGNED,
}
ASSIGNMENT_STATUS_MAP = {
    "예정": AssignmentStatus.PLANNED,
    "투입": AssignmentStatus.ASSIGNED,
    "종료": AssignmentStatus.ENDED,
    "취소": AssignmentStatus.CANCELLED,
}
ASSIGNMENT_ROLE_MAP = {
    "제안PM": ProjectAssignmentRole.PROPOSAL_PM,
    "발표PM": ProjectAssignmentRole.PRESENTATION_PM,
    "수행PM": ProjectAssignmentRole.DELIVERY_PM,
    "제안팀": ProjectAssignmentRole.PROPOSAL_TEAM,
    "수행팀": ProjectAssignmentRole.DELIVERY_TEAM,
    "지원팀": ProjectAssignmentRole.SUPPORT_TEAM,
}
BOOLEAN_TRUE = {"사용", "Y", "y", "true", "TRUE", "1", "예"}
BOOLEAN_FALSE = {"미사용", "N", "n", "false", "FALSE", "0", "아니오"}
REPRESENTATIVE_TRUE = {"Y", "y", "true", "TRUE", "1", "사용", "예"}
TABLE_MODELS = {
    "personnel": Personnel,
    "project_codes": ProjectCode,
    "projects": Project,
    "project_assignments": ProjectAssignment,
    "project_logs": ProjectLog,
    "monthly_employment_mm": MonthlyEmploymentMM,
    "monthly_assignment_mm": MonthlyAssignmentMM,
    "current_assignment_snapshots": CurrentAssignmentSnapshot,
    "weekly_load_snapshots": WeeklyLoadSnapshot,
    "monthly_kpi_summaries": MonthlyKpiSummary,
}
DELETE_ORDER = [
    MonthlyAssignmentMM,
    MonthlyEmploymentMM,
    CurrentAssignmentSnapshot,
    WeeklyLoadSnapshot,
    MonthlyKpiSummary,
    ProjectLog,
    ProjectAssignment,
    Project,
    ProjectCode,
    Personnel,
]
RESTORE_INSERT_ORDER = [
    Personnel,
    ProjectCode,
    Project,
    ProjectAssignment,
    ProjectLog,
    MonthlyEmploymentMM,
    MonthlyAssignmentMM,
    CurrentAssignmentSnapshot,
    WeeklyLoadSnapshot,
    MonthlyKpiSummary,
]
VALIDATION_CACHE: dict[str, dict[str, Any]] = {}


@dataclass
class ValidationIssue:
    sheet: str
    row_number: int
    column: str
    input_value: str
    level: str
    message: str

    def to_dict(self) -> dict[str, object]:
        return {
            "sheet": self.sheet,
            "row_number": self.row_number,
            "column": self.column,
            "input_value": self.input_value,
            "level": self.level,
            "message": self.message,
        }


def _ensure_runtime_dirs() -> None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    if not MANIFEST_PATH.exists():
        MANIFEST_PATH.write_text("[]", encoding="utf-8")


def _now() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def _require_admin(user: User) -> None:
    if user.permission != UserPermission.ADMIN:
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")


def _read_manifest() -> list[dict[str, Any]]:
    _ensure_runtime_dirs()
    try:
        return json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=500, detail="데이터 백업 manifest를 읽을 수 없습니다.") from exc


def _write_manifest(records: list[dict[str, Any]]) -> None:
    _ensure_runtime_dirs()
    MANIFEST_PATH.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def _append_manifest(record: dict[str, Any]) -> None:
    records = _read_manifest()
    records.insert(0, record)
    _write_manifest(records[:200])


def _table_count(session: Session, model: type[Any]) -> int:
    return int(session.scalar(select(func.count()).select_from(model)) or 0)


def _serialize_scalar(value: object) -> object:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if hasattr(value, "value"):
        return getattr(value, "value")
    return value


def _serialize_table_rows(session: Session, model: type[Any]) -> list[dict[str, object]]:
    rows = session.scalars(select(model)).all()
    serialized: list[dict[str, object]] = []
    for row in rows:
        payload: dict[str, object] = {}
        for column in row.__table__.columns:
            payload[column.name] = _serialize_scalar(getattr(row, column.name))
        serialized.append(payload)
    return serialized


def _make_id(prefix: str) -> str:
    stamp = _now().strftime("%Y%m%d%H%M%S")
    return f"{prefix}{stamp}{uuid4().hex[:4].upper()}"


def _dataset_counts(session: Session) -> dict[str, int]:
    return {
        "projects": _table_count(session, Project),
        "personnel": _table_count(session, Personnel),
        "assignments": _table_count(session, ProjectAssignment),
        "monthly_mm": _table_count(session, MonthlyEmploymentMM) + _table_count(session, MonthlyAssignmentMM),
        "snapshots": _table_count(session, CurrentAssignmentSnapshot) + _table_count(session, WeeklyLoadSnapshot),
    }


def _latest_backup_at(records: list[dict[str, Any]]) -> str | None:
    for record in records:
        if record.get("backup_id") and record.get("status") == "success":
            return str(record.get("created_at"))
    return None


def overview(session: Session) -> dict[str, object]:
    records = _read_manifest()
    backups = [record for record in records if record.get("backup_id")]
    restore_history = [record for record in records if str(record.get("kind", "")).startswith("restore")]
    upload_history = [record for record in records if str(record.get("kind", "")).startswith("upload")]
    return {
        "summary": {
            **_dataset_counts(session),
            "latest_backup_at": _latest_backup_at(records),
        },
        "backups": backups[:20],
        "restore_history": restore_history[:20],
        "upload_history": upload_history[:20],
    }


def _file_sha256(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def _normalize_cell(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return None if stripped in {"", "-"} else stripped
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float, Decimal)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip() or None


def _detect_header_row(rows: list[list[object | None]], required_headers: list[str]) -> int:
    for index, row in enumerate(rows[:4]):
        normalized = [(_normalize_cell(cell) or "") for cell in row]
        if all(header in normalized for header in required_headers):
            return index
    raise HTTPException(status_code=400, detail="필수 컬럼 헤더를 찾을 수 없습니다.")


def _sheet_rows(workbook: Any, sheet_name: str, required_headers: list[str]) -> tuple[list[dict[str, str | None]], list[str]]:
    if sheet_name not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail=f"필수 시트가 없습니다: {sheet_name}")
    ws = workbook[sheet_name]
    values = list(ws.iter_rows(values_only=True))
    header_row_idx = _detect_header_row(values, required_headers)
    header_values = [(_normalize_cell(cell) or "") for cell in values[header_row_idx]]
    header_positions = {header: header_values.index(header) for header in required_headers}
    records: list[dict[str, str | None]] = []
    for row_idx, row in enumerate(values[header_row_idx + 1 :], start=header_row_idx + 2):
        payload = {header: _normalize_cell(row[position] if position < len(row) else None) for header, position in header_positions.items()}
        if not any(value is not None for value in payload.values()):
            continue
        payload["_row_number"] = str(row_idx)
        records.append(payload)
    extra_columns = [header for header in header_values if header and header not in required_headers]
    return records, extra_columns


def _parse_bool(value: str | None, *, sheet: str, row_number: int, column: str, issues: list[ValidationIssue]) -> bool | None:
    if value is None:
        return None
    if value in BOOLEAN_TRUE:
        return True
    if value in BOOLEAN_FALSE:
        return False
    issues.append(ValidationIssue(sheet, row_number, column, value, "error", "지원하지 않는 사용여부 값입니다."))
    return None


def _parse_date(
    value: str | None,
    *,
    sheet: str,
    row_number: int,
    column: str,
    issues: list[ValidationIssue],
) -> date | None:
    if value is None:
        return None
    text = value.strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    issues.append(ValidationIssue(sheet, row_number, column, value, "error", "날짜 형식이 올바르지 않습니다."))
    return None


def _parse_datetime(
    value: str | None,
    *,
    sheet: str,
    row_number: int,
    column: str,
    issues: list[ValidationIssue],
) -> datetime | None:
    if value is None:
        return None
    text = value.strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y.%m.%d %H:%M", "%Y.%m.%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            if " " not in fmt:
                return datetime.combine(parsed.date(), datetime.min.time())
            return parsed
        except ValueError:
            continue
    issues.append(ValidationIssue(sheet, row_number, column, value, "error", "일시 형식이 올바르지 않습니다."))
    return None


def _parse_float(
    value: str | None,
    *,
    sheet: str,
    row_number: int,
    column: str,
    issues: list[ValidationIssue],
) -> float | None:
    if value is None:
        return None
    try:
        return float(value.replace(",", ""))
    except ValueError:
        issues.append(ValidationIssue(sheet, row_number, column, value, "error", "숫자 형식이 올바르지 않습니다."))
        return None


def _format_amount_value(value: float) -> str:
    if value >= 100_000_000:
        return f"{value / 100_000_000:.1f}억"
    if value >= 10_000:
        return f"{value / 10_000:.0f}만"
    if float(value).is_integer():
        return f"{int(value)}"
    return f"{value:.2f}"


def _parse_single_amount(value: str) -> float:
    text = value.strip().replace(",", "")
    unit_map = {"억": 100_000_000, "천만": 10_000_000, "백만": 1_000_000, "만": 10_000}
    for unit, multiplier in unit_map.items():
        if text.endswith(unit):
            return float(text[: -len(unit)]) * multiplier
    return float(text)


def _parse_amount_text(
    value: str | None,
    *,
    row_number: int,
    issues: list[ValidationIssue],
) -> tuple[float | None, float | None, str | None]:
    if value is None:
        return None, None, None
    parts = [part.strip() for part in value.split("/")]
    if len(parts) != 2 or not parts[0] or not parts[1]:
        issues.append(ValidationIssue(PROJECT_SHEET_NAME, row_number, "사업금액", value, "error", "사업금액은 총액/당사금액 형식이어야 합니다."))
        return None, None, None
    try:
        total_amount = _parse_single_amount(parts[0])
        company_amount = _parse_single_amount(parts[1])
    except ValueError:
        issues.append(ValidationIssue(PROJECT_SHEET_NAME, row_number, "사업금액", value, "error", "사업금액 형식이 올바르지 않습니다."))
        return None, None, None
    if total_amount < 0 or company_amount < 0:
        issues.append(ValidationIssue(PROJECT_SHEET_NAME, row_number, "사업금액", value, "error", "사업금액은 음수일 수 없습니다."))
        return None, None, None
    if company_amount > total_amount:
        issues.append(ValidationIssue(PROJECT_SHEET_NAME, row_number, "사업금액", value, "error", "당사 사업금액은 총 사업금액보다 클 수 없습니다."))
        return None, None, None
    return total_amount, company_amount, f"{_format_amount_value(total_amount)}/{_format_amount_value(company_amount)}"


def _coerce_mapping_enum(
    value: str | None,
    mapping: dict[str, Any],
    *,
    sheet: str,
    row_number: int,
    column: str,
    issues: list[ValidationIssue],
    required: bool = False,
) -> Any | None:
    if value is None:
        if required:
            issues.append(ValidationIssue(sheet, row_number, column, "-", "error", f"{column} 값이 필요합니다."))
        return None
    mapped = mapping.get(value)
    if mapped is None:
        issues.append(ValidationIssue(sheet, row_number, column, value, "error", f"{column} 값이 지원되지 않습니다."))
        return None
    return mapped


def _active_role_map(session: Session) -> tuple[dict[str, Role], set[str]]:
    roles = session.scalars(select(Role).where(Role.is_active.is_(True))).all()
    mapping: dict[str, Role] = {}
    duplicates: set[str] = set()
    for role in roles:
        key = (role.name or "").strip()
        if not key:
            continue
        if key in mapping:
            duplicates.add(key)
        mapping[key] = role
    return mapping, duplicates


def _validate_workbook_structure(
    workbook: Any,
    session: Session,
) -> tuple[dict[str, Any], list[ValidationIssue]]:
    issues: list[ValidationIssue] = []
    sheet_names = set(workbook.sheetnames)
    unknown_sheets = sorted(sheet_names - {PROJECT_SHEET_NAME, PERSONNEL_SHEET_NAME, ASSIGNMENT_SHEET_NAME})
    for sheet_name in unknown_sheets:
        issues.append(ValidationIssue(sheet_name, 0, "-", "-", "warning", "지원하지 않는 추가 시트는 무시됩니다."))

    project_rows, extra_project_columns = _sheet_rows(workbook, PROJECT_SHEET_NAME, PROJECT_HEADERS)
    personnel_rows, extra_personnel_columns = _sheet_rows(workbook, PERSONNEL_SHEET_NAME, PERSONNEL_HEADERS)
    assignment_rows: list[dict[str, str | None]] = []
    extra_assignment_columns: list[str] = []
    if ASSIGNMENT_SHEET_NAME in workbook.sheetnames:
        assignment_rows, extra_assignment_columns = _sheet_rows(workbook, ASSIGNMENT_SHEET_NAME, ASSIGNMENT_HEADERS)

    for column in extra_project_columns:
        issues.append(ValidationIssue(PROJECT_SHEET_NAME, 0, column, "-", "warning", "지원하지 않는 추가 컬럼은 무시됩니다."))
    for column in extra_personnel_columns:
        issues.append(ValidationIssue(PERSONNEL_SHEET_NAME, 0, column, "-", "warning", "지원하지 않는 추가 컬럼은 무시됩니다."))
    for column in extra_assignment_columns:
        issues.append(ValidationIssue(ASSIGNMENT_SHEET_NAME, 0, column, "-", "warning", "지원하지 않는 추가 컬럼은 무시됩니다."))

    if not project_rows and not personnel_rows:
        raise HTTPException(status_code=400, detail="업로드할 데이터가 없습니다.")

    active_role_by_name, duplicate_roles = _active_role_map(session)

    normalized_projects: list[dict[str, Any]] = []
    normalized_personnel: list[dict[str, Any]] = []
    normalized_assignments: list[dict[str, Any]] = []
    seen_codes: set[str] = set()
    seen_employee_numbers: set[str] = set()
    seen_emails: set[str] = set()
    seen_assignment_keys: set[str] = set()
    project_codes_in_sheet: set[str] = set()
    employee_numbers_in_sheet: set[str] = set()
    for row in project_rows:
        row_number = int(row["_row_number"] or "0")
        code = row.get("코드")
        name = row.get("사업명")
        project_type = _coerce_mapping_enum(row.get("사업유형"), PROJECT_TYPE_MAP, sheet=PROJECT_SHEET_NAME, row_number=row_number, column="사업유형", issues=issues, required=True)
        status = _coerce_mapping_enum(row.get("상태"), PROJECT_STATUS_MAP, sheet=PROJECT_SHEET_NAME, row_number=row_number, column="상태", issues=issues, required=True)
        is_active = _parse_bool(row.get("사용여부"), sheet=PROJECT_SHEET_NAME, row_number=row_number, column="사용여부", issues=issues)
        if not code:
            issues.append(ValidationIssue(PROJECT_SHEET_NAME, row_number, "코드", "-", "error", "코드는 필수입니다."))
        elif code in seen_codes:
            issues.append(ValidationIssue(PROJECT_SHEET_NAME, row_number, "코드", code, "error", "프로젝트코드는 중복될 수 없습니다."))
        else:
            seen_codes.add(code)
            project_codes_in_sheet.add(code)
        if not name:
            issues.append(ValidationIssue(PROJECT_SHEET_NAME, row_number, "사업명", "-", "error", "사업명은 필수입니다."))
        total_amount, company_amount, amount_text = _parse_amount_text(row.get("사업금액"), row_number=row_number, issues=issues)
        normalized_projects.append({
            "code": code,
            "name": name,
            "client_name": row.get("고객사"),
            "project_type": project_type.value if project_type else None,
            "status": status.value if status else None,
            "certainty": row.get("확도"),
            "amount_text": amount_text,
            "total_amount": total_amount,
            "company_amount": company_amount,
            "bid_notice_no": row.get("공고번호"),
            "bid_notice_date": _parse_date(row.get("공고일"), sheet=PROJECT_SHEET_NAME, row_number=row_number, column="공고일", issues=issues),
            "sales_department": row.get("영업부서"),
            "sales_owner": row.get("영업대표"),
            "proposal_pm_name": row.get("제안PM"),
            "presentation_pm_name": row.get("발표PM"),
            "delivery_pm_name": row.get("수행PM"),
            "start_date": _parse_date(row.get("사업 시작일"), sheet=PROJECT_SHEET_NAME, row_number=row_number, column="사업 시작일", issues=issues),
            "end_date": _parse_date(row.get("사업 종료일"), sheet=PROJECT_SHEET_NAME, row_number=row_number, column="사업 종료일", issues=issues),
            "submission_at": _parse_datetime(row.get("제안서 제출일"), sheet=PROJECT_SHEET_NAME, row_number=row_number, column="제안서 제출일", issues=issues),
            "submission_format": row.get("제출 형식"),
            "submission_note": row.get("제출 유의사항"),
            "presentation_at": _parse_datetime(row.get("제안 발표일"), sheet=PROJECT_SHEET_NAME, row_number=row_number, column="제안 발표일", issues=issues),
            "presentation_format": row.get("발표 형식"),
            "presentation_note": row.get("발표 유의사항"),
            "recent_activity_at": _parse_datetime(row.get("최근 활동일"), sheet=PROJECT_SHEET_NAME, row_number=row_number, column="최근 활동일", issues=issues),
            "is_active": is_active,
            "row_number": row_number,
        })

    for row in personnel_rows:
        row_number = int(row["_row_number"] or "0")
        employee_no = row.get("사번")
        name = row.get("성명")
        email = row.get("이메일")
        group_name = row.get("본부")
        employment_status = _coerce_mapping_enum(row.get("재직상태"), EMPLOYMENT_STATUS_MAP, sheet=PERSONNEL_SHEET_NAME, row_number=row_number, column="재직상태", issues=issues, required=True)
        is_active = _parse_bool(row.get("사용여부"), sheet=PERSONNEL_SHEET_NAME, row_number=row_number, column="사용여부", issues=issues)
        if not employee_no:
            issues.append(ValidationIssue(PERSONNEL_SHEET_NAME, row_number, "사번", "-", "error", "사번은 필수입니다."))
        elif employee_no in seen_employee_numbers:
            issues.append(ValidationIssue(PERSONNEL_SHEET_NAME, row_number, "사번", employee_no, "error", "사번은 중복될 수 없습니다."))
        else:
            seen_employee_numbers.add(employee_no)
            employee_numbers_in_sheet.add(employee_no)
        if not name:
            issues.append(ValidationIssue(PERSONNEL_SHEET_NAME, row_number, "성명", "-", "error", "성명은 필수입니다."))
        if not group_name:
            issues.append(ValidationIssue(PERSONNEL_SHEET_NAME, row_number, "본부", "-", "error", "본부는 필수입니다."))
        if email:
            if email in seen_emails:
                issues.append(ValidationIssue(PERSONNEL_SHEET_NAME, row_number, "이메일", email, "error", "이메일은 중복될 수 없습니다."))
            else:
                seen_emails.add(email)
        role_name = row.get("역할")
        role_id = None
        if role_name:
            if role_name in duplicate_roles:
                issues.append(ValidationIssue(PERSONNEL_SHEET_NAME, row_number, "역할", role_name, "error", "동일 이름의 활성 역할이 여러 개 있습니다."))
            role = active_role_by_name.get(role_name)
            if role is None:
                issues.append(ValidationIssue(PERSONNEL_SHEET_NAME, row_number, "역할", role_name, "error", "활성 역할 기준값에 없는 역할입니다."))
            else:
                role_id = role.id
        mm_start_date = _parse_date(row.get("MM 시작일"), sheet=PERSONNEL_SHEET_NAME, row_number=row_number, column="MM 시작일", issues=issues)
        mm_end_date = _parse_date(row.get("MM 종료일"), sheet=PERSONNEL_SHEET_NAME, row_number=row_number, column="MM 종료일", issues=issues)
        if mm_start_date and mm_end_date and mm_end_date < mm_start_date:
            issues.append(ValidationIssue(PERSONNEL_SHEET_NAME, row_number, "MM 종료일", row.get("MM 종료일") or "-", "error", "MM 종료일이 MM 시작일보다 빠를 수 없습니다."))
        normalized_personnel.append({
            "employee_no": employee_no,
            "name": name,
            "email": email,
            "group_name": group_name,
            "team_name": row.get("팀"),
            "position_name": row.get("직위"),
            "role_id": role_id,
            "role_name": role_name,
            "employment_status": employment_status.value if employment_status else None,
            "mm_start_date": mm_start_date,
            "mm_end_date": mm_end_date,
            "yearly_mm": _parse_float(row.get("연간 재직 MM"), sheet=PERSONNEL_SHEET_NAME, row_number=row_number, column="연간 재직 MM", issues=issues),
            "is_active": is_active,
            "row_number": row_number,
        })

    for row in assignment_rows:
        row_number = int(row["_row_number"] or "0")
        assignment_key = row.get("배정키")
        project_code = row.get("프로젝트코드")
        employee_no = row.get("사번")
        assignment_type = _coerce_mapping_enum(row.get("배정유형"), ASSIGNMENT_TYPE_MAP, sheet=ASSIGNMENT_SHEET_NAME, row_number=row_number, column="배정유형", issues=issues, required=True)
        assignment_status = _coerce_mapping_enum(row.get("배정상태"), ASSIGNMENT_STATUS_MAP, sheet=ASSIGNMENT_SHEET_NAME, row_number=row_number, column="배정상태", issues=issues) if row.get("배정상태") else None
        assignment_role = _coerce_mapping_enum(row.get("프로젝트 역할"), ASSIGNMENT_ROLE_MAP, sheet=ASSIGNMENT_SHEET_NAME, row_number=row_number, column="프로젝트 역할", issues=issues) if row.get("프로젝트 역할") else None
        if not assignment_key:
            issues.append(ValidationIssue(ASSIGNMENT_SHEET_NAME, row_number, "배정키", "-", "error", "배정키는 필수입니다."))
        elif assignment_key in seen_assignment_keys:
            issues.append(ValidationIssue(ASSIGNMENT_SHEET_NAME, row_number, "배정키", assignment_key, "error", "배정키는 중복될 수 없습니다."))
        else:
            seen_assignment_keys.add(assignment_key)
        if not project_code or project_code not in project_codes_in_sheet:
            issues.append(ValidationIssue(ASSIGNMENT_SHEET_NAME, row_number, "프로젝트코드", project_code or "-", "error", "프로젝트관리 시트의 코드와 일치해야 합니다."))
        if not employee_no or employee_no not in employee_numbers_in_sheet:
            issues.append(ValidationIssue(ASSIGNMENT_SHEET_NAME, row_number, "사번", employee_no or "-", "error", "인력관리 시트의 사번과 일치해야 합니다."))
        normalized_assignments.append({
            "assignment_key": assignment_key,
            "project_code": project_code,
            "employee_no": employee_no,
            "assignment_type": assignment_type.value if assignment_type else None,
            "assignment_role": assignment_role.value if assignment_role else None,
            "assignment_status": assignment_status.value if assignment_status else None,
            "win_loss": row.get("WIN/LOSS"),
            "onsite_type": row.get("상주유형"),
            "is_primary": (row.get("대표여부") in REPRESENTATIVE_TRUE) if row.get("대표여부") is not None else False,
            "sequence_no": _parse_float(row.get("순서"), sheet=ASSIGNMENT_SHEET_NAME, row_number=row_number, column="순서", issues=issues),
            "start_date": _parse_date(row.get("시작일"), sheet=ASSIGNMENT_SHEET_NAME, row_number=row_number, column="시작일", issues=issues),
            "end_date": _parse_date(row.get("종료일"), sheet=ASSIGNMENT_SHEET_NAME, row_number=row_number, column="종료일", issues=issues),
            "mm": _parse_float(row.get("MM"), sheet=ASSIGNMENT_SHEET_NAME, row_number=row_number, column="MM", issues=issues),
            "total_mm": _parse_float(row.get("총 MM"), sheet=ASSIGNMENT_SHEET_NAME, row_number=row_number, column="총 MM", issues=issues),
            "current_mm": _parse_float(row.get("현재 MM"), sheet=ASSIGNMENT_SHEET_NAME, row_number=row_number, column="현재 MM", issues=issues),
            "certainty_rate": _parse_float(row.get("확도율"), sheet=ASSIGNMENT_SHEET_NAME, row_number=row_number, column="확도율", issues=issues),
            "unit_price": _parse_float(row.get("단가"), sheet=ASSIGNMENT_SHEET_NAME, row_number=row_number, column="단가", issues=issues),
            "note": row.get("비고"),
            "row_number": row_number,
        })

    for assignment in normalized_assignments:
        row_number = int(assignment["row_number"])
        start_date = assignment["start_date"]
        end_date = assignment["end_date"]
        certainty_rate = assignment["certainty_rate"]
        if start_date and end_date and end_date < start_date:
            issues.append(ValidationIssue(ASSIGNMENT_SHEET_NAME, row_number, "종료일", str(end_date), "error", "종료일이 시작일보다 빠를 수 없습니다."))
        if certainty_rate is not None and not (0 <= float(certainty_rate) <= 100):
            issues.append(ValidationIssue(ASSIGNMENT_SHEET_NAME, row_number, "확도율", str(certainty_rate), "error", "확도율은 0~100 사이여야 합니다."))

    error_rows = {(issue.sheet, issue.row_number) for issue in issues if issue.level == "error" and issue.row_number > 0}
    warning_rows = {
        (issue.sheet, issue.row_number)
        for issue in issues
        if issue.level == "warning" and issue.row_number > 0 and (issue.sheet, issue.row_number) not in error_rows
    }
    total_rows = len(project_rows) + len(personnel_rows) + len(assignment_rows)
    valid_rows = max(total_rows - len(error_rows), 0)
    payload = {
        "projects": normalized_projects,
        "personnel": normalized_personnel,
        "assignments": normalized_assignments,
        "summary": {
            "sheet_count": len(workbook.sheetnames),
            "total_rows": total_rows,
            "valid_rows": valid_rows,
            "error_rows": len(error_rows),
            "warning_rows": len(warning_rows),
            "expected_counts": {
                "project_codes": len(normalized_projects),
                "projects": len(normalized_projects),
                "personnel": len(normalized_personnel),
                "project_assignments": len(normalized_assignments),
            },
        },
    }
    return payload, issues


def validate_upload(session: Session, file_name: str, raw: bytes) -> dict[str, object]:
    _ensure_runtime_dirs()
    if not file_name.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드할 수 있습니다.")
    if len(raw) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="파일 크기는 20MB 이하여야 합니다.")
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=BytesIO(raw), data_only=True)
    except ModuleNotFoundError as exc:
        raise HTTPException(status_code=503, detail="엑셀 파서(openpyxl)가 설치되어 있지 않습니다.") from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail="엑셀 파일을 읽을 수 없습니다.") from exc

    payload, issues = _validate_workbook_structure(workbook, session)
    validation_id = f"VAL-{_now().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:4].upper()}"
    file_sha = _file_sha256(raw)
    VALIDATION_CACHE[validation_id] = {
        "created_at": _now(),
        "file_sha256": file_sha,
        "file_name": file_name,
        "payload": payload,
    }
    _cleanup_validation_cache()
    return {
        "validation_id": validation_id,
        "file_name": file_name,
        "file_sha256": file_sha,
        "summary": payload["summary"],
        "issues": [issue.to_dict() for issue in issues[:200]],
    }


def _cleanup_validation_cache() -> None:
    threshold = _now() - timedelta(minutes=VALIDATION_TTL_MINUTES)
    expired = [key for key, value in VALIDATION_CACHE.items() if value["created_at"] < threshold]
    for key in expired:
        VALIDATION_CACHE.pop(key, None)


def _acquire_lock(session: Session, *, locked_by: str) -> None:
    locked_at = _now()
    lock = JobLock(
        job_name=DATA_BACKUP_JOB_NAME,
        locked_at=locked_at,
        locked_by=locked_by,
        expires_at=locked_at + timedelta(minutes=LOCK_EXPIRE_MINUTES),
    )
    try:
        session.add(lock)
        session.commit()
        return
    except IntegrityError:
        session.rollback()
    existing = session.get(JobLock, DATA_BACKUP_JOB_NAME)
    if existing is not None and existing.expires_at <= locked_at:
        session.execute(delete(JobLock).where(JobLock.job_name == DATA_BACKUP_JOB_NAME))
        session.commit()
        try:
            session.add(lock)
            session.commit()
            return
        except IntegrityError:
            session.rollback()
    raise HTTPException(status_code=409, detail="데이터 백업/업로드/복원 작업이 이미 실행 중입니다.")


def _release_lock(session: Session) -> None:
    session.execute(delete(JobLock).where(JobLock.job_name == DATA_BACKUP_JOB_NAME))
    session.commit()


def create_backup(session: Session, *, actor_name: str, kind: str, memo: str | None = None, source_file_name: str | None = None) -> dict[str, object]:
    _ensure_runtime_dirs()
    backup_id = _make_id("BKP")
    counts = {table_name: _table_count(session, model) for table_name, model in TABLE_MODELS.items()}
    payload = {
        "backup_id": backup_id,
        "created_at": _now().isoformat(timespec="seconds"),
        "actor_name": actor_name,
        "kind": kind,
        "memo": memo,
        "source_file_name": source_file_name,
        "tables": {table_name: _serialize_table_rows(session, model) for table_name, model in TABLE_MODELS.items()},
        "counts": counts,
        "included_tables": INCLUDED_TABLE_NAMES,
    }
    raw = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    file_hash = _file_sha256(raw)
    file_path = BACKUP_DIR / f"{backup_id}.json"
    file_path.write_bytes(raw)
    record = {
        "id": _make_id("OPR"),
        "created_at": payload["created_at"],
        "kind": kind,
        "status": "success",
        "backup_id": backup_id,
        "source_file_name": source_file_name,
        "actor_name": actor_name,
        "memo": memo,
        "counts": counts,
        "included_tables": INCLUDED_TABLE_NAMES,
        "file_name": file_path.name,
        "file_size": file_path.stat().st_size,
        "file_hash": file_hash,
    }
    _append_manifest(record)
    return record


def backup_detail(backup_id: str) -> dict[str, object]:
    path = BACKUP_DIR / f"{backup_id}.json"
    if not path.exists():
        raise HTTPException(status_code=404, detail="백업 파일을 찾을 수 없습니다.")
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=500, detail="백업 파일을 읽을 수 없습니다.") from exc
    return {
        "backup_id": backup_id,
        "created_at": payload.get("created_at"),
        "actor_name": payload.get("actor_name"),
        "kind": payload.get("kind"),
        "memo": payload.get("memo"),
        "source_file_name": payload.get("source_file_name"),
        "included_tables": payload.get("included_tables", INCLUDED_TABLE_NAMES),
        "counts": payload.get("counts", {}),
        "file_name": path.name,
        "file_size": path.stat().st_size,
        "file_hash": _file_sha256(path.read_bytes()),
    }


def _delete_dataset(session: Session) -> None:
    for model in DELETE_ORDER:
        session.execute(delete(model))
    session.flush()


def _relevant_months(personnel_rows: list[Personnel], assignment_rows: list[ProjectAssignment], today: date) -> list[tuple[int, int]]:
    months: set[tuple[int, int]] = {(today.year, today.month)}
    for person in personnel_rows:
        if person.mm_start_date:
            months.add((person.mm_start_date.year, person.mm_start_date.month))
        if person.mm_end_date:
            months.add((person.mm_end_date.year, person.mm_end_date.month))
    for assignment in assignment_rows:
        if assignment.start_date:
            months.add((assignment.start_date.year, assignment.start_date.month))
        if assignment.end_date:
            months.add((assignment.end_date.year, assignment.end_date.month))
    if not months:
        return [(today.year, today.month)]
    start = min(months)
    end = max(months)
    cursor = date(start[0], start[1], 1)
    end_date = date(end[0], end[1], 1)
    expanded: list[tuple[int, int]] = []
    while cursor <= end_date:
        expanded.append((cursor.year, cursor.month))
        next_month = cursor.month + 1
        next_year = cursor.year + (1 if next_month == 13 else 0)
        cursor = date(next_year, 1 if next_month == 13 else next_month, 1)
    return expanded


def _month_last_day(year: int, month: int) -> date:
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1) - timedelta(days=1)


def _person_employment_window(person: Personnel, year: int, month: int, today: date) -> tuple[date, date] | None:
    month_start = date(year, month, 1)
    month_end = _month_last_day(year, month)
    start = person.mm_start_date or month_start
    end = person.mm_end_date or month_end
    if person.employment_status in {EmploymentStatus.RETIRED, EmploymentStatus.TRANSFERRED} and person.mm_end_date is None:
        end = min(end, today)
    if end < month_start or start > month_end:
        return None
    return max(start, month_start), min(end, month_end)


def _assignment_overlap_window(assignment: ProjectAssignment, year: int, month: int) -> tuple[date, date] | None:
    if assignment.start_date is None or assignment.end_date is None:
        return None
    month_start = date(year, month, 1)
    month_end = _month_last_day(year, month)
    if assignment.end_date < month_start or assignment.start_date > month_end:
        return None
    return max(assignment.start_date, month_start), min(assignment.end_date, month_end)


def _base_assignment_mm(assignment: ProjectAssignment) -> float:
    for candidate in (assignment.current_mm, assignment.total_mm, assignment.mm):
        if candidate is not None:
            return float(candidate)
    return 0.0


def _generate_derived_rows(session: Session) -> None:
    today = date.today()
    personnel_statement = select(Personnel).outerjoin(Role, Personnel.role_id == Role.id)
    personnel_rows = session.scalars(apply_personnel_scope(personnel_statement, "pmo").where(Personnel.is_active.is_(True))).all()
    assignment_rows = session.scalars(select(ProjectAssignment)).all()
    months = _relevant_months(personnel_rows, assignment_rows, today)

    for person in personnel_rows:
        for year, month in months:
            window = _person_employment_window(person, year, month, today)
            if window is None:
                continue
            summary = month_workday_summary(session, year=year, month=month)
            workdays = int(summary["workdays"])
            employed_workdays = business_days_between(session, window[0], window[1])
            employment_mm = round((employed_workdays / workdays) if workdays else 0.0, 4)
            session.add(
                MonthlyEmploymentMM(
                    personnel_id=person.id,
                    year=year,
                    month=month,
                    workdays=workdays,
                    employed_workdays=employed_workdays,
                    employment_mm=employment_mm,
                )
            )

    for assignment in assignment_rows:
        if assignment.personnel_id is None or assignment.project_id is None:
            continue
        for year, month in months:
            window = _assignment_overlap_window(assignment, year, month)
            if window is None:
                continue
            summary = month_workday_summary(session, year=year, month=month)
            month_workdays = int(summary["workdays"])
            overlap_days = business_days_between(session, window[0], window[1])
            ratio = (overlap_days / month_workdays) if month_workdays else 0.0
            base_mm = _base_assignment_mm(assignment)
            assignment_mm = round(base_mm * ratio, 4)
            certainty_rate = float(assignment.certainty_rate) if assignment.certainty_rate is not None else None
            weighted_mm = assignment_mm * ((certainty_rate or 100.0) / 100.0)
            session.add(
                MonthlyAssignmentMM(
                    assignment_id=assignment.id,
                    project_id=assignment.project_id,
                    personnel_id=assignment.personnel_id,
                    year=year,
                    month=month,
                    assignment_mm=assignment_mm,
                    certainty_rate=certainty_rate,
                    weighted_mm=round(weighted_mm, 4),
                    assignment_type=assignment.assignment_type,
                )
            )

    active_assignments = session.scalars(select(ProjectAssignment)).all()
    project_by_id = {project.id: project for project in session.scalars(select(Project)).all()}
    for person in personnel_rows:
        person_assignments = [assignment for assignment in active_assignments if assignment.personnel_id == person.id]
        current_assignments = [
            assignment
            for assignment in person_assignments
            if assignment.start_date and assignment.end_date and assignment.start_date <= today <= assignment.end_date
        ]
        current_assignments.sort(key=lambda row: (not row.is_primary, row.start_date or today))
        next_assignments = [
            assignment for assignment in person_assignments if assignment.start_date and assignment.start_date > today
        ]
        next_assignments.sort(key=lambda row: row.start_date or today)
        current = current_assignments[0] if current_assignments else None
        next_assignment = next_assignments[0] if next_assignments else None
        representative_status = (
            current.assignment_type.value
            if current is not None
            else ("unassigned" if person.employment_status == EmploymentStatus.ACTIVE else person.employment_status.value)
        )
        current_project = project_by_id.get(current.project_id) if current and current.project_id else None
        next_project = project_by_id.get(next_assignment.project_id) if next_assignment and next_assignment.project_id else None
        session.add(
            CurrentAssignmentSnapshot(
                as_of_date=today,
                personnel_id=person.id,
                representative_status=representative_status,
                project_id=current.project_id if current else None,
                project_name=current_project.name if current_project else None,
                project_code=current_project.code if current_project else None,
                assignment_id=current.id if current else None,
                current_start_date=current.start_date if current else None,
                current_end_date=current.end_date if current else None,
                next_project_id=next_assignment.project_id if next_assignment else None,
                next_project_name=next_project.name if next_project else None,
            )
        )
        week_start = today - timedelta(days=today.weekday())
        for week_offset in range(5):
            range_start = week_start + timedelta(days=week_offset * 7)
            range_end = range_start + timedelta(days=6)
            week_assignment = next(
                (
                    assignment
                    for assignment in person_assignments
                    if assignment.start_date
                    and assignment.end_date
                    and assignment.start_date <= range_end
                    and assignment.end_date >= range_start
                ),
                None,
            )
            week_project = project_by_id.get(week_assignment.project_id) if week_assignment and week_assignment.project_id else None
            session.add(
                WeeklyLoadSnapshot(
                    as_of_date=today,
                    personnel_id=person.id,
                    week_offset=week_offset,
                    week_label=f"{range_start.month}/{range_start.day}",
                    representative_status=week_assignment.assignment_type.value if week_assignment else representative_status,
                    project_id=week_assignment.project_id if week_assignment else None,
                    project_name=week_project.name if week_project else None,
                    start_date=range_start,
                    end_date=range_end,
                )
            )

    mm_rows = session.scalars(select(MonthlyEmploymentMM)).all()
    assignment_mm_rows = session.scalars(select(MonthlyAssignmentMM)).all()
    grouped_employment: dict[tuple[int, int], float] = defaultdict(float)
    grouped_assignment: dict[tuple[int, int, str], float] = defaultdict(float)
    for row in mm_rows:
        grouped_employment[(row.year, row.month)] += float(row.employment_mm or 0)
    for row in assignment_mm_rows:
        grouped_assignment[(row.year, row.month, row.assignment_type.value if row.assignment_type else "unassigned")] += float(
            row.weighted_mm if row.weighted_mm is not None else row.assignment_mm
        )
    months_for_kpi = sorted({(row.year, row.month) for row in mm_rows} | {(row.year, row.month) for row in assignment_mm_rows})
    for year, month in months_for_kpi:
        avg_headcount_mm = grouped_employment[(year, month)]
        running_mm = grouped_assignment[(year, month, AssignmentType.DELIVERY.value)]
        proposing_mm = grouped_assignment[(year, month, AssignmentType.PROPOSAL.value)]
        support_mm = grouped_assignment[(year, month, AssignmentType.SUPPORT.value)]
        assigned_total = running_mm + proposing_mm + support_mm
        idle_mm = max(avg_headcount_mm - assigned_total, 0.0)
        utilization = round((running_mm / avg_headcount_mm) * 100, 2) if avg_headcount_mm else 0.0
        contract = round((assigned_total / avg_headcount_mm) * 100, 2) if avg_headcount_mm else 0.0
        session.add(
            MonthlyKpiSummary(
                year=year,
                month=month,
                organization_name="PMO본부",
                avg_headcount_mm=round(avg_headcount_mm, 4),
                running_mm=round(running_mm, 4),
                proposing_mm=round(proposing_mm, 4),
                support_mm=round(support_mm, 4),
                idle_mm=round(idle_mm, 4),
                utilization_rate=utilization,
                contract_rate=contract,
                source_snapshot_date=today,
            )
        )


def _insert_validated_payload(session: Session, payload: dict[str, Any], actor_name: str) -> None:
    personnel_id_by_employee_no: dict[str, str] = {}
    project_code_id_by_code: dict[str, str] = {}
    project_id_by_code: dict[str, str] = {}

    for row in payload["personnel"]:
        person = Personnel(
            employee_no=row["employee_no"],
            name=row["name"],
            email=row["email"],
            group_name=row["group_name"],
            team_name=row["team_name"],
            position_name=row["position_name"],
            role_id=row["role_id"],
            role_name=row["role_name"],
            employment_status=EmploymentStatus(row["employment_status"]),
            mm_start_date=row["mm_start_date"],
            mm_end_date=row["mm_end_date"],
            yearly_mm=row["yearly_mm"],
            is_active=bool(row["is_active"]) if row["is_active"] is not None else True,
        )
        session.add(person)
        session.flush()
        if person.employee_no:
            personnel_id_by_employee_no[person.employee_no] = person.id

    for row in payload["projects"]:
        code = row["code"]
        project_code = ProjectCode(
            code=code,
            name=row["name"],
            project_type=ProjectType(row["project_type"]),
            status=ProjectStatus(row["status"]),
            certainty=row["certainty"],
            is_active=bool(row["is_active"]) if row["is_active"] is not None else True,
            source_sheet=PROJECT_SHEET_NAME,
        )
        session.add(project_code)
        session.flush()
        project = Project(
            project_code_id=project_code.id,
            code=code,
            name=row["name"],
            client_name=row["client_name"],
            sales_department=row["sales_department"],
            sales_owner=row["sales_owner"],
            project_type=ProjectType(row["project_type"]),
            status=ProjectStatus(row["status"]),
            certainty=row["certainty"],
            proposal_pm_name=row["proposal_pm_name"],
            presentation_pm_name=row["presentation_pm_name"],
            delivery_pm_name=row["delivery_pm_name"],
            amount_text=row["amount_text"],
            total_amount=row["total_amount"],
            company_amount=row["company_amount"],
            start_date=row["start_date"],
            end_date=row["end_date"],
            bid_notice_no=row["bid_notice_no"],
            bid_notice_date=row["bid_notice_date"],
            submission_at=row["submission_at"],
            submission_format=row["submission_format"],
            submission_note=row["submission_note"],
            presentation_at=row["presentation_at"],
            presentation_format=row["presentation_format"],
            presentation_note=row["presentation_note"],
            recent_activity_at=row["recent_activity_at"],
        )
        session.add(project)
        session.flush()
        project_code_id_by_code[code] = project_code.id
        project_id_by_code[code] = project.id

    for row in payload["assignments"]:
        project_id = project_id_by_code.get(row["project_code"])
        personnel_id = personnel_id_by_employee_no.get(row["employee_no"])
        assignment = ProjectAssignment(
            project_id=project_id,
            personnel_id=personnel_id,
            assignment_type=AssignmentType(row["assignment_type"]),
            assignment_role=ProjectAssignmentRole(row["assignment_role"]) if row["assignment_role"] else None,
            assignment_status=AssignmentStatus(row["assignment_status"]) if row["assignment_status"] else None,
            win_loss=row["win_loss"],
            onsite_type=row["onsite_type"],
            is_primary=bool(row["is_primary"]),
            sequence_no=int(row["sequence_no"]) if row["sequence_no"] is not None else None,
            start_date=row["start_date"],
            end_date=row["end_date"],
            mm=row["mm"],
            total_mm=row["total_mm"],
            current_mm=row["current_mm"],
            certainty_rate=row["certainty_rate"],
            unit_price=row["unit_price"],
            note=row["note"],
            source_sheet=ASSIGNMENT_SHEET_NAME,
        )
        session.add(assignment)

    session.flush()
    _generate_derived_rows(session)
    session.flush()
def apply_validated_upload(session: Session, *, validation_id: str, memo: str | None, user: User) -> dict[str, object]:
    _require_admin(user)
    cached = VALIDATION_CACHE.get(validation_id)
    if cached is None:
        raise HTTPException(status_code=404, detail="검증 결과를 찾을 수 없습니다. 다시 검증해 주세요.")
    if cached["created_at"] < _now() - timedelta(minutes=VALIDATION_TTL_MINUTES):
        VALIDATION_CACHE.pop(validation_id, None)
        raise HTTPException(status_code=409, detail="검증 결과가 만료되었습니다. 다시 검증해 주세요.")
    summary = cached["payload"]["summary"]
    if summary["error_rows"] > 0:
        raise HTTPException(status_code=409, detail="검증 오류가 있는 파일은 반영할 수 없습니다.")

    actor_name = user.name or "관리자"
    pre_backup = create_backup(
        session,
        actor_name=actor_name,
        kind="upload_before_backup",
        memo=memo,
        source_file_name=cached["file_name"],
    )
    _acquire_lock(session, locked_by=actor_name)
    try:
        _delete_dataset(session)
        _insert_validated_payload(session, cached["payload"], actor_name)
        session.commit()
    except Exception as exc:  # noqa: BLE001
        session.rollback()
        _append_manifest(
            {
                "id": _make_id("OPR"),
                "created_at": _now().isoformat(timespec="seconds"),
                "kind": "upload_failed",
                "status": "failed",
                "backup_id": None,
                "source_file_name": cached["file_name"],
                "actor_name": actor_name,
                "memo": memo,
                "message": str(exc),
            }
        )
        raise
    finally:
        _release_lock(session)
    record = {
        "id": _make_id("OPR"),
        "created_at": _now().isoformat(timespec="seconds"),
        "kind": "upload_complete",
        "status": "success",
        "backup_id": None,
        "source_file_name": cached["file_name"],
        "actor_name": actor_name,
        "memo": memo,
        "validation_id": validation_id,
        "file_sha256": cached["file_sha256"],
        "counts": _dataset_counts(session),
        "pre_backup_id": pre_backup["backup_id"],
    }
    _append_manifest(record)
    VALIDATION_CACHE.pop(validation_id, None)
    return record


def _load_backup_payload(backup_id: str) -> dict[str, Any]:
    path = BACKUP_DIR / f"{backup_id}.json"
    if not path.exists():
        raise HTTPException(status_code=404, detail="복원할 백업을 찾을 수 없습니다.")
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=500, detail="백업 파일이 손상되었습니다.") from exc


def _deserialize_scalar(value: Any, column: Any) -> Any:
    if value is None:
        return None
    column_type = column.type
    if isinstance(column_type, DateTime):
        return datetime.fromisoformat(value)
    if isinstance(column_type, Date):
        return date.fromisoformat(value)
    if isinstance(column_type, Enum) and getattr(column_type, "enum_class", None) is not None:
        return column_type.enum_class(value)
    if isinstance(column_type, Numeric):
        return Decimal(str(value))
    if isinstance(column_type, Integer):
        return int(value)
    return value


def _deserialize_row(model: type[Any], payload: dict[str, Any]) -> dict[str, Any]:
    next_payload: dict[str, Any] = {}
    for column in model.__table__.columns:
        if column.name not in payload:
            continue
        next_payload[column.name] = _deserialize_scalar(payload[column.name], column)
    return next_payload


def _restore_table(session: Session, model: type[Any], rows: list[dict[str, Any]]) -> None:
    for payload in rows:
        session.add(model(**_deserialize_row(model, payload)))
    session.flush()


def restore_backup(session: Session, *, backup_id: str, memo: str | None, user: User) -> dict[str, object]:
    _require_admin(user)
    backup_payload = _load_backup_payload(backup_id)
    actor_name = user.name or "관리자"
    current_backup = create_backup(session, actor_name=actor_name, kind="restore_before_backup", memo=memo)
    role_ids = {role.id for role in session.scalars(select(Role).where(Role.is_active.is_(True))).all()}
    for row in backup_payload.get("tables", {}).get("personnel", []):
        role_id = row.get("role_id")
        if role_id and role_id not in role_ids:
            raise HTTPException(status_code=409, detail="백업 데이터가 현재 활성 역할 기준값과 맞지 않아 복원할 수 없습니다.")
    _acquire_lock(session, locked_by=actor_name)
    try:
        _delete_dataset(session)
        for model in RESTORE_INSERT_ORDER:
            table_name = model.__tablename__
            _restore_table(session, model, backup_payload.get("tables", {}).get(table_name, []))
        session.commit()
    except Exception as exc:  # noqa: BLE001
        session.rollback()
        _append_manifest(
            {
                "id": _make_id("OPR"),
                "created_at": _now().isoformat(timespec="seconds"),
                "kind": "restore_failed",
                "status": "failed",
                "backup_id": None,
                "actor_name": actor_name,
                "memo": memo,
                "source_backup_id": backup_id,
                "message": str(exc),
            }
        )
        raise
    finally:
        _release_lock(session)
    record = {
        "id": _make_id("OPR"),
        "created_at": _now().isoformat(timespec="seconds"),
        "kind": "restore_complete",
        "status": "success",
        "backup_id": None,
        "actor_name": actor_name,
        "memo": memo,
        "source_backup_id": backup_id,
        "counts": _dataset_counts(session),
        "pre_backup_id": current_backup["backup_id"],
    }
    _append_manifest(record)
    return record
